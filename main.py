import requests
import os
import shutil
import zipfile
import csv
import traceback

import xlrd
from openpyxl import load_workbook

import config

def write_file(file_name, data):
  with open(file_name, 'wb') as fw:
    for chunk in data.iter_content(chunk_size=512 * 1024): 
      if chunk: 
        fw.write(chunk)


def download_files():
  try:
    index = config.START_YEAR
    while index >= config.END_YEAR:
      if config.START_YEAR - index < 2:
        month = 12
        while month > 0:
          response = requests.get(config.DOWNLOAD_XLSX_URL % (index, month))
          if response.status_code == 200:
            write_file(config.XSLX_FILE_NAME % (index, month), response)
          month -= 1
      else:
        response = requests.get(config.DOWNLOAD_ZIP_URL % (index))
        if response.status_code == 200:
          write_file(config.ZIP_FILE_NAME % (index), response)
      index-= 1
    return True
  except:
    print(traceback.format_exc())
    return False


def extract_from_zip_files():
  try:
    index = config.START_YEAR -2
    while index >= config.END_YEAR:
      with zipfile.ZipFile(config.ZIP_FILE_NAME % index) as z:
        month = 12
        while month > 0:
          try:
            with z.open(config.EXTRACT_FILE_NAME[index] % (index, index, month)) as zf, open(config.XSLX_FILE_NAME % (index, month), 'wb') as f:
              shutil.copyfileobj(zf, f)
          except:
            print('Error extracting file',config.XSLX_FILE_NAME % (index, month))
            pass
          month -= 1
      index -= 1
    return True
  except:
    print(traceback.format_exc())
    return False


def process_files():
  try:
    records = []
    for file in os.listdir(config.INPUT_FOLDER):
      if file.startswith('table7'):
        wb = xlrd.open_workbook(config.INPUT_FOLDER + file)
        sheet = wb.sheet_by_index(0) 
        split_date = (file.replace('.xlsx','')).split('-')[1]

        for i in range(sheet.nrows):
          if sheet.cell_value(i, 0) in [0,1]:
            records.append({
              "Year": split_date[0:4],
              "Month": split_date.replace(split_date[0:4],''),
              "Expenditure category": sheet.cell_value(i, 1),
              "Relative importance": sheet.cell_value(i,2),
              "Unadjusted percent change": sheet.cell_value(i,3),
              "Unadjusted effect on All Items": sheet.cell_value(i,4)
            })
    if len(records) > 0:
      with open(config.PROCESSED_DATA, 'w', newline='') as fw:
        dw = csv.DictWriter(fw,fieldnames=records[0].keys())
        dw.writeheader()
        for row in records:
          dw.writerow(row)
        return records
    else:
      return None
  except:
    print(traceback.format_exc())
    return None


def update_excel(data):
  try:
    wb = load_workbook(filename = config.EXCEL_FILE, read_only=False, keep_vba=True)
    ws = wb[config.EXCEL_TAB] 
    row_number = 2
    for row in data:
      ws.cell(row=row_number, column=1, value=int(row['Year']))
      ws.cell(row=row_number, column=2, value=int(row['Month']))
      ws.cell(row=row_number, column=3, value=row['Expenditure category'])
      ws.cell(row=row_number, column=4, value=row['Relative importance'])
      ws.cell(row=row_number, column=5, value=row['Unadjusted percent change'])
      ws.cell(row=row_number, column=6, value=row['Unadjusted effect on All Items'])
      row_number += 1
    wb.save(config.EXCEL_FILE)
    return True
  except:
    print(traceback.format_exc())
    return False


def main():
  if download_files():
    print('Files downloaded...')
    if extract_from_zip_files():
      print('Files extracted from zip files...')
      raw_data = process_files()
      if raw_data:
        print('Data extracted from files and formatted...')
        if update_excel(raw_data):
          print("Process Executed Successfully.")
      else:
        print("Error processing XSLX files.")
    else:
      print("Error extracting XSLX files.")
  else:
    print("Error downloading files.")

if __name__ == '__main__':
  main()